#!/usr/bin/env python3
"""
课表发送脚本 - 读取Excel并生成格式化消息
支持发送到Pushplus微信推送服务
"""

import os
from datetime import datetime
from openpyxl import load_workbook
import requests

# ====== 配置区域 ======
EXCEL_PATH = os.environ.get("EXCEL_PATH", "schedule.xlsx")
PUSHPLUS_TOKEN = os.environ.get("PUSHPLUS_TOKEN", "3cc0c70f320746b2ab4a6ef7a144fa4f")
PUSHPLUS_TOPIC = os.environ.get("PUSHPLUS_TOPIC", "")
LOCATION = os.environ.get("LOCATION", "港汇店")

# ====== 函数定义 ======
def get_weekday_chinese():
    weekdays = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
    return weekdays[datetime.now().weekday()]

def get_date_str():
    now = datetime.now()
    return f"💭{now.year}.{now.month}.{now.day}  {get_weekday_chinese()}"

def read_schedule_from_excel(file_path):
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        schedule_data = []
        for row_num in range(4, 9):
            row = ws[row_num]
            time_cell = row[1].value
            if not time_cell:
                continue
            time_str = str(time_cell).strip()
            if '-' in time_str:
                simple_time = time_str.split('-')[0].strip()
            else:
                simple_time = time_str
            for col_idx in range(2, 9):
                weekday_map = {2: "周一", 3: "周二", 4: "周三", 5: "周四", 6: "周五", 7: "周六", 8: "周日"}
                weekday = weekday_map.get(col_idx)
                if not weekday:
                    continue
                course_cell = row[col_idx].value
                if course_cell:
                    schedule_data.append({'weekday': weekday, 'time': simple_time, 'raw_content': str(course_cell)})
        print(f"✅ 成功读取Excel文件，共 {len(schedule_data)} 条课程")
        return schedule_data
    except Exception as e:
        print(f"❌ 读取Excel文件失败: {e}")
        return None

def parse_course_info(course_str):
    if not course_str:
        return None, None
    course_str = str(course_str).strip()
    if not course_str:
        return None, None
    parts = course_str.replace('\\n', '\n').split('\n')
    parts = [p.strip() for p in parts if p.strip()]
    if len(parts) >= 3:
        coach = parts[-1]
        full_name = ' · '.join(parts[:-1])
        return coach, full_name
    elif len(parts) == 2:
        return parts[1], parts[0] + ' · ' + parts[1]
    elif len(parts) == 1:
        return '', parts[0]
    return None, None

def format_schedule_message(schedule_data, weekday, location="港汇店"):
    date_str = get_date_str()
    message_parts = [f"{date_str} /{location}", ""]
    today_courses = [c for c in schedule_data if c['weekday'] == weekday]
    if not today_courses:
        message_parts.append("今日暂无课程安排")
    else:
        for course in today_courses:
            time_str = course['time']
            coach, full_name = parse_course_info(course['raw_content'])
            if coach and full_name:
                course_entry = f"🌸{time_str} {coach}【{full_name}】"
                message_parts.append(course_entry)
    message_parts.extend([
        "",
        "❣️ 提前5分钟进入教室",
        "❣️手机静音暂存前台或更衣柜",
        "❣️当天约课即可",
        "❣️约课时间：早7:30－晚20:00",
        f"地址：港汇中心3楼"
    ])
    return "\n".join(message_parts)

def send_to_pushplus(message, token, topic=""):
    url = "https://www.pushplus.plus/send"
    if topic:
        url = "https://www.pushplus.plus/topic/send"
        data = {"token": token, "topic": topic, "content": message, "type": "text"}
    else:
        data = {"token": token, "content": message, "type": "text"}
    try:
        response = requests.post(url, json=data, timeout=10)
        result = response.json()
        if result.get("code") == 200:
            print("✅ 消息发送成功！")
            return True
        else:
            print(f"❌ 发送失败: {result.get('msg', '未知错误')}")
            return False
    except Exception as e:
        print(f"❌ 发送异常: {e}")
        return False

def main():
    print("=" * 50)
    print("🧘 课表发送脚本")
    print("=" * 50)
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ Excel文件不存在: {EXCEL_PATH}")
        return
    weekday = get_weekday_chinese()
    print(f"📅 今天是：{weekday}")
    schedule_data = read_schedule_from_excel(EXCEL_PATH)
    if schedule_data is None:
        return
    message = format_schedule_message(schedule_data, weekday, LOCATION)
    print("\n📋 消息预览：")
    print("-" * 50)
    print(message)
    print("-" * 50)
    print("\n📤 正在发送消息...")
    success = send_to_pushplus(message, PUSHPLUS_TOKEN, PUSHPLUS_TOPIC)
    if success:
        print("\n🎉 完成！请检查微信接收消息。")
    else:
        print("\n⚠️ 发送可能失败，请检查Token和群编码配置。")

if __name__ == "__main__":
    main()
